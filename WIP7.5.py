mr = True
while mr == True:
    ##Application Completion through Esign
    y = input('Enter a number of records to process: ')
    rn = input('Enter row number to start from: ')
    x = 1
    rn1 = int(rn) 
    while x <= int(y): #specify number of times to run process
        import os
        from datetime import datetime, date
        from selenium import webdriver
        from selenium.webdriver.common.by import By
        from selenium.webdriver.common.keys import Keys
        from selenium.webdriver.support.ui import WebDriverWait
        from selenium.webdriver.support import expected_conditions as EC
        from selenium.common.exceptions import NoSuchElementException
        from selenium.common.exceptions import ElementNotInteractableException
        from openpyxl import Workbook
        from openpyxl import load_workbook
        wb1 = load_workbook('WebAppData1.xlsx')
        ws1 = wb1['AppDataList']
        error = (NoSuchElementException, ElementNotInteractableException)
        ## Following col variables are workbook/sheet cell locations based on letter and specified row num (rn1) above
        col1 = 'C'+str(rn1)     #PIP First Name
        col2 = 'D'+str(rn1)     #PIP Last Name
        col3 = 'E'+str(rn1)     #PIP User ID
        col4 = 'F'+str(rn1)     #PIP SSN
        col5 = 'G'+str(rn1)     #PIP DOB
        col6 = 'H'+str(rn1)     #PIP Gender
        col7 = 'I'+str(rn1)    #Child First Name
        col8 = 'J'+str(rn1)    #Child DOB
        col9 = 'K'+str(rn1)    #Child Gender
        col10 = 'L'+str(rn1)    #City
        col11 = 'M'+str(rn1)    #zip
        col12 = 'N'+str(rn1)    #PIP Race
        col13 = 'O'+str(rn1)   #Child Race
        col14 = 'P'+str(rn1)    #Has Income?
        col15 = 'Q'+str(rn1)    #Income Amount
        col16 = 'R'+str(rn1)    #Pay Freq
        col17 = 'S'+str(rn1)    #Number of Hours
        col18 = 'T'+str(rn1)    #Start Date
        col19 = 'U'+str(rn1)   #Unearned Income?
        col20 = 'V'+str(rn1)   #Unearned Amount
        col21 = 'W'+str(rn1)   #Unearned Start Date
        col22 = 'X'+str(rn1)    #Apply for FA
        col23 = 'Y'+str(rn1)    #Apply for TCA
        col24 = 'Z'+str(rn1)    #Apply for Medicaid
        col25 = 'AA'+str(rn1)    ##Apply for SSI Related Medicaid
        ## Following variables are workbook/sheet cell variables
        fn1 = ws1[col1]     #PIP First Name
        ln1 = ws1[col2]     #PIP Last Name
        ui1 = ws1[col3]     #PIP User ID
        sn1 = ws1[col4]     #PIP SSN
        db1 = ws1[col5]     #PIP DOB
        gd1 = ws1[col6]     #PIP Gender
        cfn1 = ws1[col7]    #Child First Name
        cdb1 = ws1[col8]    #Child DOB
        cgd1 = ws1[col9]    #Child Gender
        ct1 = ws1[col10]    #City
        zp1 = ws1[col11]    #zip
        rc1 = ws1[col12]    #PIP Race
        crc1 = ws1[col13]   #Child Race
        hi1 = ws1[col14]    #Has Income?
        ia1 = ws1[col15]    #Income Amount
        fq1 = ws1[col16]    #Pay Freq
        nh1 = ws1[col17]    #Number of Hours
        id1 = ws1[col18]    #Start Date
        uni1 = ws1[col19]   #Unearned Income?
        una1 = ws1[col20]   #Unearned Amount
        und1 = ws1[col21]   #Unearned Start Date
        af1 = ws1[col22]    #Apply for FA
        at1 = ws1[col23]    #Apply for TCA
        am1 = ws1[col24]    #Apply for MediCaid
        am2 = ws1[col25]    #Apply for SSI Related Medicaid
        ## Following variables are specific cell values 
        if fn1.value is None:
            print("No Data in this row.")
            break
        else:
            fnv1 = fn1.value

            
        lnv1 = ln1.value
        uiv1 = ui1.value
        snv1 = sn1.value
        if db1.value is None:
            pass
        else:
            dbv1 = db1.value.strftime("%m%d%Y")


        if gd1.value is None:
            gdv1 = input('Enter the Gender of {}: '.format(fnv1)).upper()
        else:
            gdv1 = gd1.value.upper()


        cfnv1 = cfn1.value
        if cdb1.value is None:
            pass
        else:
            cdbv1 = cdb1.value.strftime("%m%d%Y")


        if cgd1.value is None:
            cgdv1 = input('Enter the Gender of {}: '.format(cfnv1)).upper()
        else:
            cgdv1 = cgd1.value.upper()



        ctv1 = ct1.value
        zpv1 = zp1.value
        if rc1.value is None:
            pass
        else:
            rcv1 = rc1.value.upper()


        if crc1.value is None:
            pass
        else:
            crcv1 = crc1.value.upper()


        if hi1.value is None:
            pass
        else:
            hiv1 = hi1.value.upper()


        iav1 = ia1.value
        if fq1.value is None:
            fqv1 = fq1.value
        else:
            fqv1 = fq1.value.upper()


        nhv1 = nh1.value
        if id1.value is None or id1.value == "":
            pass
        else:
            idv1 = id1.value.strftime("%m%d%Y")


        if uni1.value is None:
            pass
        else:
            univ1 = uni1.value.upper()


        unav1 = una1.value
        if und1.value is None or und1.value == "":
            pass
        else:
            undv1 = und1.value.strftime("%m%d%Y")


        if af1.value is None:
            afv1 = 'N'
        else:
            afv1 = af1.value.upper()


        if at1.value is None:
            atv1 = 'N'
        else:
            atv1 = at1.value.upper()


        if am1.value is None:
            amv1 = 'N'
        else:
            amv1 = am1.value.upper()


        if am2.value is None:
            amv2 = 'N'
        else:
            amv2 = am2.value.upper()

        browser = webdriver.Firefox(executable_path=r'C:\Users\garren-james\AppData\Local\geckodriver')
        browser.get('http://160.131.243.164/access/index.do')
        ##Start of application completion
        ##User Account Setup
        htmlElem = browser.find_element_by_tag_name('html')
        browser.maximize_window()
        linkElem = browser.find_element_by_link_text('Apply for Benefits')
        type(linkElem)
        linkElem.click()
        htmlElem = WebDriverWait(browser, 60).until(EC.presence_of_element_located((By.ID, 'startNewApplicationRadio')))
        htmlElem = browser.find_element_by_tag_name('#startNewApplicationRadio')
        htmlElem.click()
        htmlElem = browser.find_element_by_name('imgNext')
        htmlElem.click()
        htmlElem = WebDriverWait(browser, 60).until(EC.presence_of_element_located((By.ID, 'createAccount')))
        htmlElem = browser.find_element_by_tag_name('#createAccount')
        htmlElem.click()
        htmlElem = browser.find_element_by_name('imgNext')
        htmlElem.click()
        htmlElem = WebDriverWait(browser, 60).until(EC.presence_of_element_located((By.ID, 'firstName')))
        htmlElem = browser.find_element_by_tag_name('#firstName')
        htmlElem.send_keys(fnv1)
        htmlElem = browser.find_element_by_tag_name('#lastName')
        htmlElem.send_keys(lnv1)
        if len(uiv1) >= 6:
            htmlElem = browser.find_element_by_tag_name('#userId')
            htmlElem.send_keys(uiv1)


        if len(uiv1) < 6:
            htmlElem = browser.find_element_by_tag_name('#userId')
            htmlElem.send_keys(uiv1+'123')


        htmlElem = browser.find_element_by_tag_name('#password')
        htmlElem.send_keys('test123')
        htmlElem = browser.find_element_by_tag_name('#retypePassword')
        htmlElem.send_keys('test123')
        htmlElem = browser.find_element_by_tag_name('#securityQuestion1')
        htmlElem.send_keys(Keys.DOWN)
        htmlElem = browser.find_element_by_tag_name('#answer1')
        htmlElem.send_keys('aa')
        htmlElem = browser.find_element_by_tag_name('#securityQuestion2')
        htmlElem.send_keys(Keys.DOWN)
        htmlElem = browser.find_element_by_tag_name('#answer2')
        htmlElem.send_keys('aa')
        htmlElem = browser.find_element_by_tag_name('#securityQuestion3')
        htmlElem.send_keys(Keys.DOWN)
        htmlElem = browser.find_element_by_tag_name('#answer3')
        htmlElem.send_keys('aa')
        checkbox = browser.find_element_by_css_selector("input#userAgreement")
        checkbox.click()
        browser.execute_script("window.scrollTo(0, document.documentElement.scrollHeight);")
        htmlElem = browser.find_element_by_name('imgNext')
        htmlElem.click()
        htmlElem = browser.find_element_by_class_name('footerButton')
        htmlElem.click()
        htmlElem = WebDriverWait(browser, 60).until(EC.presence_of_element_located((By.ID, 'userId')))
        if len(uiv1) >= 6:
            htmlElem = browser.find_element_by_tag_name('#userId')
            htmlElem.send_keys(uiv1)


        if len(uiv1) < 6:
            htmlElem = browser.find_element_by_tag_name('#userId')
            htmlElem.send_keys(uiv1+'123')


        htmlElem = browser.find_element_by_tag_name('#password')
        htmlElem.send_keys('test123')
        htmlElem = browser.find_element_by_class_name('signInButtonEN')
        htmlElem.click()
        htmlElem = browser.find_element_by_name('imgNext')
        htmlElem.click()
        ##Start of application questions/demographics
        htmlElem = WebDriverWait(browser, 60).until(EC.presence_of_element_located((By.ID, 'myselfAndFamily')))
        htmlElem = browser.find_element_by_tag_name('#myselfAndFamily')
        htmlElem.click()
        htmlElem = browser.find_element_by_name('imgNext')
        htmlElem.click()
        ##Benefits Applied for Selections
        htmlElem = WebDriverWait(browser, 60).until(EC.presence_of_element_located((By.ID, 'foodStampsCheckbox')))
        if afv1 == 'Y':
            checkbox = browser.find_element_by_css_selector("input#foodStampsCheckbox")
            checkbox.click()
        else:
                pass


        if atv1 == 'Y':
            checkbox = browser.find_element_by_css_selector("input#cashAssistanceCheckbox1")
            checkbox.click()
        else:
                pass


        if amv1 == 'Y':
            checkbox = browser.find_element_by_css_selector("input#acaMedicalAssitanceCheckbox")
            checkbox.click()
        else:
                pass


        if amv2 == 'Y':
            checkbox = browser.find_element_by_css_selector("input#medicalAssitanceCheckbox")
            checkbox.click()
        else:
                pass


        htmlElem = browser.find_element_by_name('imgNext')
        htmlElem.click()
        htmlElem = WebDriverWait(browser, 3)
        if afv1 == 'Y' or atv1 == 'Y':
            htmlElem = browser.find_element_by_name('imgNext')
            htmlElem.click()
        else:
            pass

            
        ##PIP Demo Information + Address
        htmlElem = WebDriverWait(browser, 60).until(EC.presence_of_element_located((By.ID, 'firstName')))
        htmlElem = browser.find_element_by_tag_name('#firstName')
        htmlElem.send_keys(fnv1)
        htmlElem = browser.find_element_by_tag_name('#lastName')
        htmlElem.send_keys(lnv1)
        if gdv1 == 'F':
            htmlElem = browser.find_element_by_tag_name('#genderFemaleRadio')
            htmlElem.click()


        if gdv1 == 'M':
            htmlElem = browser.find_element_by_tag_name('#genderMaleRadio')
            htmlElem.click()


        htmlElem = browser.find_element_by_tag_name('#dob')
        htmlElem.send_keys(dbv1)
        htmlElem = browser.find_element_by_tag_name('#address1')
        htmlElem.send_keys('1000 Anywhere St')
        htmlElem = browser.find_element_by_tag_name('#city')
        htmlElem.send_keys(ctv1)
        htmlElem = browser.find_element_by_tag_name('#zipCode')
        htmlElem.send_keys(zpv1)
        htmlElem = browser.find_element_by_tag_name('#notSameAsLivingAddress')
        htmlElem.click()
        htmlElem = browser.find_element_by_name('imgNext')
        htmlElem.click()
        ##Code 1 Validation Pop Up
        htmlElem = WebDriverWait(browser, 1)
        def is_add_val_present():
            try:
                htmlElem = browser.find_element_by_tag_name('#addressSelection1')
                htmlElem.click()
                htmlElem = browser.find_element_by_name('imgNext')
                htmlElem.click()
            except NoSuchElementException:
                pass
            

        is_add_val_present()
        htmlElem = browser.find_element_by_name('imgNext')
        htmlElem.click()
        htmlElem = WebDriverWait(browser, 60).until(EC.presence_of_element_located((By.ID, 'noticeLanguageE')))
        htmlElem = browser.find_element_by_tag_name('#noticeLanguageE')
        htmlElem.click()
        htmlElem = browser.find_element_by_tag_name('#maritalStatus')
        htmlElem.send_keys(Keys.DOWN)
        htmlElem.send_keys(Keys.DOWN)
        htmlElem.send_keys(Keys.DOWN)
        htmlElem.send_keys(Keys.DOWN)
        htmlElem.send_keys(Keys.DOWN)
        htmlElem.send_keys(Keys.DOWN)
        htmlElem = browser.find_element_by_tag_name('#livingArrangmentType')
        htmlElem.send_keys(Keys.DOWN)
        def tax_filer_present():
            try:
                htmlElem = browser.find_element_by_tag_name('#taxFilerYes')
                htmlElem.click()
            except NoSuchElementException:
                pass
            

        tax_filer_present()
        htmlElem = browser.find_element_by_tag_name('#ssn')
        htmlElem.send_keys(snv1)
        htmlElem = browser.find_element_by_tag_name('#indvAliasNameSsnNo')
        htmlElem.click()
        htmlElem = browser.find_element_by_tag_name('#raceAsApplied')
        if rcv1 == 'I':
            htmlElem.send_keys(Keys.DOWN)
            htmlElem = browser.find_element_by_tag_name('#ferderalN')
            htmlElem.click()


        if rcv1 == 'A':
            htmlElem.send_keys(Keys.DOWN)
            htmlElem.send_keys(Keys.DOWN)


        if rcv1 == 'B':
            htmlElem.send_keys(Keys.DOWN)
            htmlElem.send_keys(Keys.DOWN)
            htmlElem.send_keys(Keys.DOWN)


        if rcv1 == 'T':
            htmlElem.send_keys(Keys.DOWN)
            htmlElem.send_keys(Keys.DOWN)
            htmlElem.send_keys(Keys.DOWN)
            htmlElem.send_keys(Keys.DOWN)
            htmlElem.send_keys(Keys.DOWN)


        if rcv1 == 'W':
            htmlElem.send_keys(Keys.DOWN)
            htmlElem.send_keys(Keys.DOWN)
            htmlElem.send_keys(Keys.DOWN)
            htmlElem.send_keys(Keys.DOWN)
            htmlElem.send_keys(Keys.DOWN)
            htmlElem.send_keys(Keys.DOWN)
            htmlElem.send_keys(Keys.DOWN)
            htmlElem.send_keys(Keys.DOWN)


        htmlElem = browser.find_element_by_tag_name('#appliesForBenefitsYes')
        htmlElem.click()
        htmlElem = browser.find_element_by_tag_name('#addindividualYes')
        htmlElem.click()
        htmlElem = browser.find_element_by_name('imgNext')
        htmlElem.click()
        ##Add Code to address known SSNs here
        ##PIP Citizenship Page
        htmlElem = WebDriverWait(browser, 60).until(EC.presence_of_element_located((By.ID, 'floridaResidentYes')))
        def place_born_present():
            try:
                htmlElem = browser.find_element_by_tag_name('#placeOfBirth')
                htmlElem.send_keys(Keys.DOWN)
            except NoSuchElementException:
                pass


        place_born_present()
        htmlElem = browser.find_element_by_tag_name('#floridaResidentYes')
        htmlElem.click()
        def disabled():
            try:
                if amv2 == 'Y':
                    htmlElem = browser.find_element_by_tag_name('#disabilityYes')
                    htmlElem.click()
                elif amv2 == 'N':
                    htmlElem = browser.find_element_by_tag_name('#disabilityNo')
                    htmlElem.click()
            except NoSuchElementException:
                pass
            

        disabled()
        htmlElem = browser.find_element_by_tag_name('#indvCitizenshipYes')
        htmlElem.click()
        def out_US_present():
            try:
                htmlElem = browser.find_element_by_tag_name('#outOfUSNo')
                htmlElem.click()
            except NoSuchElementException:
                pass


        out_US_present()
        htmlElem = browser.find_element_by_name('imgNext')
        htmlElem.click()
        ##Child Demo Information
        htmlElem = WebDriverWait(browser, 60).until(EC.presence_of_element_located((By.ID, 'firstName')))
        htmlElem = browser.find_element_by_tag_name('#firstName')
        htmlElem.send_keys(cfnv1)
        htmlElem = browser.find_element_by_tag_name('#lastName')
        htmlElem.send_keys(lnv1)
        if cgdv1 == 'F':
            htmlElem = browser.find_element_by_tag_name('#indvSexFemale')
            htmlElem.click()


        if cgdv1 == 'M':
            htmlElem = browser.find_element_by_tag_name('#indvSexMale')
            htmlElem.click()


        htmlElem = browser.find_element_by_tag_name('#dateOfBirth')
        htmlElem.send_keys(cdbv1)
        htmlElem = browser.find_element_by_tag_name('#maritalStatus')
        htmlElem.send_keys(Keys.DOWN)
        htmlElem.send_keys(Keys.DOWN)
        htmlElem.send_keys(Keys.DOWN)
        htmlElem.send_keys(Keys.DOWN)
        htmlElem.send_keys(Keys.DOWN)
        htmlElem.send_keys(Keys.DOWN)
        htmlElem = browser.find_element_by_tag_name('#livingArrangmentType')
        htmlElem.send_keys(Keys.DOWN)
        def tax_child_present():
            try:
                htmlElem = browser.find_element_by_tag_name('#taxFilerNo')
                htmlElem.click()
            except NoSuchElementException:
                pass


        tax_child_present()
        htmlElem = browser.find_element_by_tag_name('#ssn')
        htmlElem.send_keys(snv1+1)
        htmlElem = browser.find_element_by_tag_name('#indvAliasNameSsnNo')
        htmlElem.click()
        htmlElem = browser.find_element_by_tag_name('#raceAsApplied')
        if rcv1 == 'I':
            htmlElem.send_keys(Keys.DOWN)
            htmlElem = browser.find_element_by_tag_name('#ferderalN')
            htmlElem.click()


        if rcv1 == 'A':
            htmlElem.send_keys(Keys.DOWN)
            htmlElem.send_keys(Keys.DOWN)


        if rcv1 == 'B':
            htmlElem.send_keys(Keys.DOWN)
            htmlElem.send_keys(Keys.DOWN)
            htmlElem.send_keys(Keys.DOWN)


        if rcv1 == 'T':
            htmlElem.send_keys(Keys.DOWN)
            htmlElem.send_keys(Keys.DOWN)
            htmlElem.send_keys(Keys.DOWN)
            htmlElem.send_keys(Keys.DOWN)
            htmlElem.send_keys(Keys.DOWN)


        if rcv1 == 'W':
            htmlElem.send_keys(Keys.DOWN)
            htmlElem.send_keys(Keys.DOWN)
            htmlElem.send_keys(Keys.DOWN)
            htmlElem.send_keys(Keys.DOWN)
            htmlElem.send_keys(Keys.DOWN)
            htmlElem.send_keys(Keys.DOWN)
            htmlElem.send_keys(Keys.DOWN)
            htmlElem.send_keys(Keys.DOWN)


        htmlElem = browser.find_element_by_tag_name('#appliesForBenefitsYes')
        htmlElem.click()
        htmlElem = browser.find_element_by_tag_name('#addindividualNo')
        htmlElem.click()
        htmlElem = WebDriverWait(browser, 3)
        htmlElem = browser.find_element_by_name('imgNext')
        htmlElem.click()
        htmlElem = browser.find_element_by_name('imgNext')
        htmlElem.click()
        ##Child Citizenship
        htmlElem = WebDriverWait(browser, 60).until(EC.presence_of_element_located((By.ID, 'floridaResidentYes')))
        place_born_present()
        htmlElem = browser.find_element_by_tag_name('#floridaResidentYes')
        htmlElem.click()
        def cld_disab_present():
            try:
                htmlElem = browser.find_element_by_tag_name('#disabilityNo')
                htmlElem.click()
            except NoSuchElementException:
                pass
            

        cld_disab_present()
        htmlElem = browser.find_element_by_tag_name('#indvCitizenshipYes')
        htmlElem.click()
        out_US_present()
        htmlElem = browser.find_element_by_name('imgNext')
        htmlElem.click()
        ##Rights Resp+Hippa
        htmlElem = WebDriverWait(browser, 60).until(EC.presence_of_element_located((By.ID, 'rightsResp')))
        htmlElem = browser.find_element_by_tag_name('#rightsResp')
        htmlElem.click()
        htmlElem = browser.find_element_by_name('imgNext')
        htmlElem.click()
        def is_hipaa_present():
            try:
                htmlElem = browser.find_element_by_tag_name('#rightsHipaa')
                htmlElem.click()
                htmlElem = browser.find_element_by_name('imgNext')
                htmlElem.click()
            except NoSuchElementException:
                pass
            

        is_hipaa_present()
        ##Individual Relationship Section
        htmlElem = WebDriverWait(browser, 60).until(EC.presence_of_element_located((By.NAME, 'relationship[0].relation')))
        if gdv1 == 'F':
            htmlElem = browser.find_element_by_name('relationship[0].relation')
            htmlElem.send_keys(Keys.DOWN)
            htmlElem.send_keys(Keys.DOWN)
            htmlElem.send_keys(Keys.DOWN)
            htmlElem.send_keys(Keys.DOWN)
            htmlElem.send_keys(Keys.DOWN)
            htmlElem.send_keys(Keys.DOWN)
            htmlElem.send_keys(Keys.DOWN)
            htmlElem.send_keys(Keys.DOWN)
            htmlElem.send_keys(Keys.DOWN)
            

        if gdv1 == 'M':
            htmlElem = browser.find_element_by_name('relationship[0].relation')
            htmlElem.send_keys(Keys.DOWN)
            htmlElem.send_keys(Keys.DOWN)
            htmlElem.send_keys(Keys.DOWN)


        def purc_prep():
            try:
                htmlElem = browser.find_element_by_name('relationship[0].purchaseOrPrepareMealsTogether')
                htmlElem.click()
                htmlElem = browser.find_element_by_name('imgNext')
                htmlElem.click()
            except error:
                htmlElem = browser.find_element_by_name('imgNext')
                htmlElem.click()


        def is_ooth_present():
            try:
                htmlElem = browser.find_element_by_name('dependents')
                htmlElem.click()
                htmlElem = browser.find_element_by_tag_name('#oothNo')
                htmlElem.click()
                htmlElem = browser.find_element_by_name('imgNext')
                htmlElem.click()
            except NoSuchElementException:
                pass

            
        purc_prep()
        is_ooth_present()
        ##Pregnancy Page
        def is_preg_present():
            try:
                htmlElem = browser.find_element_by_tag_name('#pregnancyInfoInHome0')
                htmlElem.click()
                htmlElem = browser.find_element_by_name('imgNext')
                htmlElem.click()
            except NoSuchElementException:
                pass
            

        is_preg_present()
        ##Other HH Info Question Page
        htmlElem = WebDriverWait(browser, 60).until(EC.presence_of_element_located((By.ID, 'multipleStateBenefitsIndvs')))
        def is_renal_present():
            try:
                htmlElem = browser.find_element_by_tag_name('#renalDialysis0')
                htmlElem.click()
            except NoSuchElementException:
                pass
            

        def is_school_ind_present():
            try:
                htmlElem = browser.find_element_by_tag_name('#schoolIndvs')
                htmlElem.click()
            except NoSuchElementException:
                pass
            

        def is_flee_fln_present():
            try:
                htmlElem = browser.find_element_by_tag_name('#fleeingLawIndvs')
                htmlElem.click()
            except NoSuchElementException:
                pass
            

        def is_drug_trfk_present():
            try:
                htmlElem = browser.find_element_by_tag_name('#drugTrafickingFelonyIndvs')
                htmlElem.click()
            except NoSuchElementException:
                pass
            

        def is_other_state_present():
            try:
                htmlElem = browser.find_element_by_tag_name('#benefitOtherStateIndvs0')
                htmlElem.click()
            except NoSuchElementException:
                pass
            

        is_renal_present()
        is_school_ind_present()
        is_flee_fln_present()
        is_drug_trfk_present()
        htmlElem = browser.find_element_by_tag_name('#multipleStateBenefitsIndvs')
        htmlElem.click()
        is_other_state_present()
        def is_ssi_present():
            try:
                htmlElem = browser.find_element_by_tag_name('#SsiIndvs')
                htmlElem.click()
            except NoSuchElementException:
                pass


        def is_fcare_present():
            try:
                htmlElem = browser.find_element_by_tag_name('#priorFosterCare0')
                htmlElem.click()
            except NoSuchElementException:
                pass


        def is_daily_living():
            try:
                htmlElem = browser.find_element_by_tag_name('#dailyLivingAssistance0')
                htmlElem.click()
            except NoSuchElementException:
                pass


        def is_chld_limit():
            try:
                htmlElem = browser.find_element_by_tag_name('#ChildLimitedIndvs0')
                htmlElem.click()
            except NoSuchElementException:
                pass


        def is_chld_ther():
            try:
                htmlElem = browser.find_element_by_tag_name('#ChildTherapy0')
                htmlElem.click()
            except NoSuchElementException:
                pass


        def is_chld_medC():
            try:
                htmlElem = browser.find_element_by_tag_name('#ChildMedicare0')
                htmlElem.click()
            except NoSuchElementException:
                pass


        def is_immune_present():
            try:
                htmlElem = browser.find_element_by_tag_name('#ImmunizationIndv0')
                htmlElem.click()
            except NoSuchElementException:
                pass


        def is_hlth_check():
            try:
                htmlElem = browser.find_element_by_tag_name('#healthCheckupIndvs0')
                htmlElem.click()
            except NoSuchElementException:
                pass


        def is_adj():
            try:
                htmlElem = browser.find_element_by_tag_name('#AdultbyJudge0')
                htmlElem.click()
            except NoSuchElementException:
                pass


        def is_fmr_fc():
            try:
                htmlElem = browser.find_element_by_tag_name('#fosterChildIndvs0')
                htmlElem.click()
            except NoSuchElementException:
                pass


        def is_mgt_frmwk():
            try:
                htmlElem = browser.find_element_by_tag_name('#migrantOrFarmWorkerNo')
                htmlElem.click()
            except NoSuchElementException:
                pass


        is_ssi_present()
        is_fcare_present()
        is_daily_living()
        is_chld_limit()
        is_chld_ther()
        is_chld_medC()
        is_immune_present()
        is_hlth_check()
        is_adj()
        is_fmr_fc()
        is_mgt_frmwk()
        htmlElem = browser.find_element_by_name('imgNext')
        htmlElem.click()
        ##Certify identity for children
        def is_cert_present():
            try:
                browser.execute_script("window.scrollTo(0, document.body.scrollHeight);")
                htmlElem = browser.find_element_by_name('performAction')
                htmlElem.click()
            except error:
                pass


        is_cert_present()
        ##Absent Parent Details
        def is_AP_present():
            try:
                htmlElem = browser.find_element_by_tag_name('#fnameId')
                htmlElem.send_keys('unknown')
                htmlElem = browser.find_element_by_tag_name('#lnameId')
                htmlElem.send_keys('unknown')
                htmlElem = browser.find_element_by_tag_name('#AbsParentSex')
                if gdv1 == 'F':
                    htmlElem.send_keys(Keys.DOWN)
                    htmlElem.send_keys(Keys.DOWN)


                if gdv1 == 'M':
                    htmlElem.send_keys(Keys.DOWN)


                htmlElem = browser.find_element_by_tag_name('#LegalParentYes')
                htmlElem.click()
                htmlElem = browser.find_element_by_name('parentOf')
                htmlElem.click()
                htmlElem = browser.find_element_by_tag_name('#EnforcementServiceYes')
                htmlElem.click()
                htmlElem = browser.find_element_by_tag_name('#ReasonCode')
                htmlElem.send_keys(Keys.DOWN)
                htmlElem.send_keys(Keys.DOWN)
                htmlElem.send_keys(Keys.DOWN)
                htmlElem.send_keys(Keys.DOWN)
                htmlElem.send_keys(Keys.DOWN)
                htmlElem.send_keys(Keys.DOWN)
                htmlElem.send_keys(Keys.DOWN)
                htmlElem = browser.find_element_by_tag_name('#AnotherParentNo')
                htmlElem.click()
                htmlElem = browser.find_element_by_name('imgNext')
                htmlElem.click()
            except NoSuchElementException:
                pass


        is_AP_present()
        ##Lifeline Page
        def is_lifeline_present():
            try:
                htmlElem = browser.find_element_by_tag_name('#caseReducedPhoneChoiceNo')
                htmlElem.click()
                htmlElem = browser.find_element_by_name('imgNext')
                htmlElem.click()
            except NoSuchElementException:
                pass
            finally:
                htmlElem = browser.find_element_by_name('imgNext')
                htmlElem.click()



        is_lifeline_present()
        ##Assets
        def is_cash_indv():
            try:
                htmlElem = browser.find_element_by_tag_name('#CashIndvs0')
                htmlElem.click()
            except NoSuchElementException:
                pass


        def is_bank_indv():
            try:
                htmlElem = browser.find_element_by_tag_name('#BankIndvs0')
                htmlElem.click()
            except NoSuchElementException:
                pass


        def is_other_indv():
            try:
                htmlElem = browser.find_element_by_tag_name('#OtherIndvs0')
                htmlElem.click()
            except NoSuchElementException:
                pass


        def is_transf_indv():
            try:
                htmlElem = browser.find_element_by_tag_name('#TransferredIndvs0')
                htmlElem.click()
            except NoSuchElementException:
                pass


        def is_cash_settle():
            try:
                htmlElem = browser.find_element_by_tag_name('#CashSettlementIndvs0')
                htmlElem.click()
            except NoSuchElementException:
                pass


        is_cash_indv()
        is_bank_indv()
        is_other_indv()
        is_transf_indv()
        is_cash_settle()
        htmlElem = browser.find_element_by_name('imgNext')
        htmlElem.click()
        htmlElem = browser.find_element_by_name('imgNext')
        htmlElem.click()
        ##Assets Cont
        def is_life_ins():
            try:
                htmlElem = browser.find_element_by_tag_name('#LifeInsurance1')
                htmlElem.click()
            except NoSuchElementException:
                pass


        def is_car():
            try:
                htmlElem = browser.find_element_by_tag_name('#Vehicle1')
                htmlElem.click()
            except NoSuchElementException:
                pass


        def is_real_prop():
            try:
                htmlElem = browser.find_element_by_tag_name('#realpro')
                htmlElem.click()
            except NoSuchElementException:
                pass


        def is_buss_asset():
            try:
                htmlElem = browser.find_element_by_tag_name('#businessasset1')
                htmlElem.click()
                htmlElem = browser.find_element_by_name('imgNext')
                htmlElem.click()
                htmlElem = browser.find_element_by_name('imgNext')
                htmlElem.click()
            except NoSuchElementException:
                pass


        is_life_ins()
        is_car()
        is_real_prop()
        is_buss_asset()
        ##Earned Income
        htmlElem = WebDriverWait(browser, 60).until(EC.presence_of_element_located((By.ID, 'employment0')))
        if hiv1 == 'N':
            htmlElem = browser.find_element_by_tag_name('#employment0')
            htmlElem.click()


        if hiv1 == 'Y':
            htmlElem = browser.find_element_by_class_name('imageBoxDiv')
            htmlElem.click()


        htmlElem = browser.find_element_by_tag_name('#incomeEndLast')
        htmlElem.click()
        htmlElem = browser.find_element_by_tag_name('#selfEmployment0')
        htmlElem.click()
        htmlElem = browser.find_element_by_tag_name('#roomAndBoardIncome0')
        htmlElem.click()
        def is_refuse_wrk():
            try:
                htmlElem = browser.find_element_by_tag_name('#jobOfferAndRefused0')
                htmlElem.click()
            except NoSuchElementException:
                pass


        def is_on_strike():
            try:
                htmlElem = browser.find_element_by_tag_name('#anyoneOnStrike0')
                htmlElem.click()
            except NoSuchElementException:
                pass


        is_refuse_wrk()
        is_on_strike()
        htmlElem = browser.find_element_by_name('imgNext')
        htmlElem.click()
        def add_employ():
            htmlElem = browser.find_element_by_tag_name('#employerName')
            htmlElem.send_keys('Generic Employer')
            htmlElem = browser.find_element_by_tag_name('#jobBeginDate')
            htmlElem.send_keys(idv1)
            htmlElem = browser.find_element_by_tag_name('#payFrequency')  
            if fqv1 == 'B':
                htmlElem.send_keys(Keys.DOWN)
            if fqv1 == 'M':
                htmlElem.send_keys(Keys.DOWN)
                htmlElem.send_keys(Keys.DOWN)
            if fqv1 == 'T':
                htmlElem.send_keys(Keys.DOWN)
                htmlElem.send_keys(Keys.DOWN)
                htmlElem.send_keys(Keys.DOWN)
                htmlElem.send_keys(Keys.DOWN)
            if fqv1 == 'W':
                htmlElem.send_keys(Keys.DOWN)
                htmlElem.send_keys(Keys.DOWN)
                htmlElem.send_keys(Keys.DOWN)
                htmlElem.send_keys(Keys.DOWN)
                htmlElem.send_keys(Keys.DOWN)
            htmlElem = browser.find_element_by_tag_name('#monthlyHours')
            htmlElem.send_keys(nhv1)
            htmlElem = browser.find_element_by_tag_name('#monthlyHHIPAmount')
            htmlElem.send_keys(iav1)
            htmlElem = browser.find_element_by_tag_name('#anotherEmploymentNo')
            htmlElem.click()
            htmlElem = browser.find_element_by_name('imgNext')
            htmlElem.click()


        if hiv1 == 'Y':
            add_employ()

        if hiv1 == 'N':
            pass

        htmlElem = browser.find_element_by_name('imgNext')
        htmlElem.click()
        ##Unearned Income
        htmlElem = WebDriverWait(browser, 60).until(EC.presence_of_element_located((By.ID, 'alimony0')))
        def chld_supt_present():
            try:
                htmlElem = browser.find_element_by_tag_name('#childsupport0')
                htmlElem.click()
            except NoSuchElementException:
                pass


        chld_supt_present()
        htmlElem = browser.find_element_by_tag_name('#alimony0')
        htmlElem.click()
        def ssi_present():
            try:
                htmlElem = browser.find_element_by_tag_name('#supplsecurityincome0')
                htmlElem.click()
            except NoSuchElementException:
                pass


        ssi_present()
        htmlElem = browser.find_element_by_tag_name('#ssincome0')
        htmlElem.click()
        def ind_inc_present():
            try:
                htmlElem = browser.find_element_by_tag_name('#scrflalaskannativeinfo0')
                htmlElem.click()
            except NoSuchElementException:
                pass


        ind_inc_present()
        if univ1 == 'N':
            htmlElem = browser.find_element_by_tag_name('#otherTypes0')
            htmlElem.click()

        if univ1 == 'Y':
            htmlElem = browser.find_elements_by_name('otherTypes')[1]
            htmlElem.click()


        def unern_present():
            try:
                htmlElem = browser.find_element_by_tag_name('#scrflunearnedincomeapplied0')
                htmlElem.click()
            except NoSuchElementException:
                pass


        def educa_aid_present():
            try:
                htmlElem = browser.find_element_by_tag_name('#educationAid0')
                htmlElem.click()
            except NoSuchElementException:
                pass


        def tax_deduct_present():
            try:
                htmlElem = browser.find_element_by_tag_name('#taxDeductions0')
                htmlElem.click()
            except NoSuchElementException:
                pass


        unern_present()
        educa_aid_present()
        tax_deduct_present()
        htmlElem = browser.find_element_by_name('imgNext')
        htmlElem.click()
        ##Adding Unearned Income
        def add_unearn():
            htmlElem = browser.find_element_by_name('individuals[0].mapTypes(OT)')
            htmlElem.click()
            htmlElem = browser.find_element_by_name('imgNext')
            htmlElem.click()
            htmlElem = browser.find_element_by_tag_name('#beginDateId')
            htmlElem.send_keys(undv1)
            htmlElem = browser.find_element_by_tag_name('#amt')
            htmlElem.send_keys(unav1)
            htmlElem = browser.find_element_by_tag_name('#subTypeId')
            htmlElem.send_keys(Keys.DOWN)
            htmlElem.send_keys(Keys.DOWN)
            htmlElem.send_keys(Keys.DOWN)
            htmlElem.send_keys(Keys.DOWN)
            htmlElem.send_keys(Keys.DOWN)
            htmlElem = browser.find_element_by_tag_name('#nextincomeN')
            htmlElem.click()
            htmlElem = browser.find_element_by_name('imgNext')
            htmlElem.click()


        if univ1 == 'Y':
            add_unearn()

        if univ1 == 'N':
            pass


        htmlElem = browser.find_element_by_name('imgNext')
        htmlElem.click()
        ##HH Expenses
        def house_expns_present():
            try:
                htmlElem = browser.find_element_by_tag_name('#shelter0')
                htmlElem.click()
                htmlElem = browser.find_element_by_tag_name('#utility0')
                htmlElem.click()
                htmlElem = browser.find_element_by_tag_name('#roomBoardExpenses0')
                htmlElem.click()
                htmlElem = browser.find_element_by_tag_name('#homeless0')
                htmlElem.click()
                htmlElem = browser.find_element_by_tag_name('#heatcool0')
                htmlElem.click()
                htmlElem = browser.find_element_by_tag_name('#receiveLiheapNo')
                htmlElem.click()
                htmlElem = browser.find_element_by_name('imgNext')
                htmlElem.click()
                htmlElem = browser.find_element_by_name('imgNext')
                htmlElem.click()
            except NoSuchElementException:
                pass


        def TCA_expns_present():
            try:
                htmlElem = browser.find_element_by_tag_name('#shelter0')
                htmlElem.click()
                htmlElem = browser.find_element_by_tag_name('#roomBoardExpenses0')
                htmlElem.click()
                htmlElem = browser.find_element_by_name('imgNext')
                htmlElem.click()
                htmlElem = browser.find_element_by_name('imgNext')
                htmlElem.click()
            except NoSuchElementException:
                pass


        if afv1 == 'N':
            TCA_expns_present()
        else:
            house_expns_present()


        ##Other Expenses
        def cldsupt_exp_present():
            try:
                htmlElem = browser.find_element_by_tag_name('#childSupport0')
                htmlElem.click()
            except NoSuchElementException:
                pass


        def dep_care_present():
            try:
                htmlElem = browser.find_element_by_tag_name('#childAdultDayCare0')
                htmlElem.click()
            except NoSuchElementException:
                pass


        def med_exp_present():
            try:
                htmlElem = browser.find_element_by_tag_name('#ongoingMedicalBills0')
                htmlElem.click()
            except NoSuchElementException:
                pass


        def retro_med_present():
            try:
                htmlElem = browser.find_element_by_tag_name('#unpaidMedicalBills0')
                htmlElem.click()
            except NoSuchElementException:
                pass


        def medicare_present():
            try:
                htmlElem = browser.find_element_by_tag_name('#medicare0')
                htmlElem.click()
            except NoSuchElementException:
                pass


        def blind_work_present():
            try:
                htmlElem = browser.find_element_by_tag_name('#blindWithWorkRelatedExpenses0')
                htmlElem.click()
            except NoSuchElementException:
                pass


        def hlth_ins_present():
            try:
                htmlElem = browser.find_element_by_tag_name('#scrflHealthInsuranceNo')
                htmlElem.click()
            except NoSuchElementException:
                pass


        def empl_hlth_ins():
            try:
                htmlElem = browser.find_element_by_tag_name('#scrflDeclinedHealthInsuranceNo')
                htmlElem.click()
            except NoSuchElementException:
                pass


        def volcan_hlth_ins():
            try:
                htmlElem = browser.find_element_by_tag_name('#voluntarycancelhlthnsflagNo')
                htmlElem.click()
            except NoSuchElementException:
                pass


        cldsupt_exp_present()
        dep_care_present()
        med_exp_present()
        retro_med_present()
        medicare_present()
        blind_work_present()
        hlth_ins_present()
        empl_hlth_ins()
        volcan_hlth_ins()
        htmlElem = browser.find_element_by_name('imgNext')
        htmlElem.click()
        htmlElem = browser.find_element_by_name('imgNext')
        htmlElem.click()
        htmlElem = browser.find_element_by_name('imgNext')
        htmlElem.click()
        htmlElem = browser.find_element_by_name('imgNext')
        htmlElem.click()
        ##SOU
        htmlElem = WebDriverWait(browser, 60).until(EC.presence_of_element_located((By.ID, 'reviewed')))
        htmlElem = browser.find_element_by_tag_name('#reviewed')
        htmlElem.click()
        htmlElem = browser.find_element_by_name('imgNext')
        htmlElem.click()
        if afv1 == 'N' and atv1 == 'N':
            htmlElem = browser.find_element_by_name('imgNext')
            htmlElem.click()
        else:
            pass


        ##Lexis/Nexis questions
        if afv1 == 'Y' or atv1 == 'Y':
            htmlElem = WebDriverWait(browser, 60).until(EC.presence_of_element_located((By.NAME, 'question[0].selectedanswer')))
            htmlElem = browser.find_element_by_name('question[0].selectedanswer')
            htmlElem.click()
            htmlElem = browser.find_element_by_name('question[1].selectedanswer')
            htmlElem.click()
            htmlElem = browser.find_element_by_name('question[2].selectedanswer')
            htmlElem.click()
            htmlElem = browser.find_element_by_name('question[3].selectedanswer')
            htmlElem.click()
            htmlElem = browser.find_element_by_name('imgNext')
            htmlElem.click()
        else:
            htmlElem = WebDriverWait(browser, 60).until(EC.presence_of_element_located((By.NAME, 'farsVerification')))
            htmlElem = browser.find_element_by_tag_name('#farsVerificationP')
            htmlElem.click()
            htmlElem = browser.find_element_by_name('imgNext')
            htmlElem.click()
            

        ##Esign Section
        htmlElem = WebDriverWait(browser, 3)
        htmlElem = browser.find_element_by_name('imgNext')
        htmlElem.click()
        htmlElem = browser.find_element_by_tag_name('#caseVoteChoiceNo')
        htmlElem.click()
        htmlElem = browser.find_element_by_tag_name('#esigncheckbox')
        htmlElem.click()
        htmlElem = browser.find_element_by_name('imgNext')
        htmlElem.click()
        ##Application Wrap Up
        htmlElem = WebDriverWait(browser, 60).until(EC.presence_of_element_located((By.ID, 'surveyN')))
        htmlElem = browser.find_element_by_tag_name('#surveyN')
        htmlElem.click()
        htmlElem = browser.find_element_by_tag_name('#emailFlagN')
        htmlElem.click()
        htmlElem = browser.find_element_by_name('imgNext')
        htmlElem.click()
        htmlElem = browser.find_element_by_name('imgSaveExit')
        htmlElem.click()
        ##Iteration controls
        browser.quit()
        rn1 = rn1+1
        x = x+1
    c = input ('Process other rows? (Y or N): ')
    if c.upper() == 'Y':
        mr = True
    else:
        mr = False
        an = True
        ##Start of WIP Number Lookup
        while an == True:
            c = input ('Look up ACCESS/Case Numbers? (Y or N): ')
            if c.upper() == 'Y':
                y = input('Enter a number of records to get ACCESS/Case Numbers for: ')
                rn = input('Enter row number to start from: ')
                x = 1
                rn1 = int(rn)
                while x <= int(y):
                    import os
                    from datetime import datetime, date
                    from selenium import webdriver
                    from selenium.webdriver.common.by import By
                    from selenium.webdriver.common.keys import Keys
                    from selenium.webdriver.support.ui import WebDriverWait
                    from selenium.webdriver.support import expected_conditions as EC
                    from selenium.common.exceptions import NoSuchElementException
                    from openpyxl import Workbook
                    from openpyxl import load_workbook
                    wb1 = load_workbook('WebAppData1.xlsx')
                    ws1 = wb1.active
                    col3 = 'E'+str(rn1)     ##UserID
                    col4 = 'A'+str(rn1)     ##ACCESS Number
                    col5 = 'B'+str(rn1)    ##Case Number
                    ui1 = ws1[col3]
                    an1 = ws1[col4]
                    uiv1 = ui1.value
                    anv1 = an1.value
                    if ui1.value is None:
                        print("No Data in this row.")
                        break
                    else:
                        uiv1 = ui1.value

                
                    browser = webdriver.Firefox(executable_path=r'C:\Users\garren-james\AppData\Local\geckodriver')
                    browser.get('http://160.131.243.164/access/index.do')
                    htmlElem = browser.find_element_by_tag_name('#userId')
                    htmlElem.send_keys(uiv1)
                    htmlElem = browser.find_element_by_tag_name('#password')
                    htmlElem.send_keys('test123')
                    htmlElem = browser.find_element_by_class_name('signInButtonEN')
                    htmlElem.click()
                    table_element = browser.find_element_by_class_name('fieldset_outer_table')
                    cellloc = table_element.find_element_by_xpath('table/tbody/tr[2]/td[2]')
                    cellloc = cellloc.text
                    cellloc = int(cellloc)
                    ws1[col4] = cellloc
                    wb1.save('WebAppData1.xlsx')
                    col4 = 'A'+str(rn1)     ##\
                    an1 = ws1[col4]         ## |Reload the Variable with the new data in it
                    anv1 = an1.value        ##/
                    browser.get('http://ams-accp.dcf.state.fl.us/AMSR2/')
                    htmlElem = browser.find_element_by_name('loginform:userid')
                    htmlElem.send_keys('fz0383')
                    htmlElem = browser.find_element_by_name('loginform:password')
                    htmlElem.send_keys('moxie123')
                    htmlElem = browser.find_element_by_name('loginform:logonbutton')
                    htmlElem.click()
                    htmlElem = browser.find_element_by_name('homeform:header1:case1')
                    htmlElem.send_keys(anv1)
                    htmlElem = browser.find_element_by_name('homeform:header1:Gobutton')
                    htmlElem.click()
                    try:
                        table_element = browser.find_element_by_class_name('myFaces_panelTabbedPane')
                        cell = table_element.find_element_by_xpath('.//tr[1]/td[6]')
                        casenum = cell.get_attribute('innerHTML')
                        casenum = casenum[102:112]
                        casenum = int(casenum)
                        ws1[col5] = casenum
                        wb1.save('WebAppData1.xlsx')
                    except NoSuchElementException:
                        browser.quit()
                        print('TIP has not processed, try again in a few minutes.')
                        break
                    browser.quit()
                    rn1 = rn1+1
                    x = x+1
                    ##Some stuff
                c = input ('Process other rows? (Y or N): ')
                if c.upper() == 'Y':
                    an = True
                else:
                    an = False
                